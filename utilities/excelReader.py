import os
import pathlib

import openpyxl


# cell = sheet.cell(row=2, column=1)
# print(cell.value)

# sheet.cell(row=5, column=1).value = "Samsung"  # write data in sheet
# print(sheet.max_row)
# print(sheet.max_column)
# print(sheet['A3'].value)
#
# for row in range(1, sheet.max_row + 1):
#     for column in range(1, sheet.max_column + 1):
#         print(sheet.cell(row=row, column=column).value)
#     print("------------------")
def open_excel(filename, sheetname= None):
    cwd = pathlib.Path.cwd()
    dir = os.path.dirname(cwd)
    excel_path = os.path.join(cwd, "testData", filename)
    print(excel_path)
    workbook = openpyxl.load_workbook(excel_path)
    return workbook[sheetname] if sheetname is not None else workbook.active

@staticmethod
def get_excel_data(filename, sheetname= None):
    sheet = open_excel(filename, sheetname)
    list = []
    for row in range(2, sheet.max_row + 1):
        dict = {}
        for column in range(1, sheet.max_column + 1):
            dict[sheet.cell(1, column).value] = sheet.cell(row, column).value
        list.append(dict)
    return list


print(get_excel_data("testData.xlsx", "testSheet2"))
